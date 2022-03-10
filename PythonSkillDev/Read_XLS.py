from openpyxl.chart import BubbleChart, Reference, Series
from openpyxl.chart import BarChart, LineChart, Reference
import openpyxl
import xlrd


# COMMAND TO CHECK INSTALLED PYTHON PACKAGES
# pip list
# pip install xlrd==1.2.0

workbook = xlrd.open_workbook('excel.xlsx')

excel_sheet = workbook.sheet_by_index(0)
excel_sheet.cell_value(0, 0)

for i in range(excel_sheet.ncols):
    print(excel_sheet.cell_value(i, 0),
          excel_sheet.cell_value(i, 1),
          excel_sheet.cell_value(i, 2))


wb_obj = openpyxl.load_workbook('excel.xlsx')
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=2)

print(cell_obj.value)

row = sheet_obj.max_row
column = sheet_obj.max_column


print(row, column)


for rw in range(1, row + 1):
    for col in range(1, column + 1):
        cell_obj = sheet_obj.cell(row=rw, column=col)
        print(cell_obj.value)


cell_obj = sheet_obj['A1':'B7']

for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)


wb = openpyxl.Workbook()

sheet = wb.active

for i in range(10):
    sheet.append([i])


values = Reference(sheet, min_col=1,
                   min_row=1,
                   max_col=1,
                   max_row=10)

chart = LineChart()
chart.add_data(values)
chart.title = "BAR CHART"

chart.x_axis.title = "X_AXIS"
chart.y_axis.title = "Y_AXIS"

sheet.add_chart(chart, 'E2')

wb.save('sample.xlsx')


wb = openpyxl.Workbook()

sheet = wb.active

rows = [
    ("Number of Products", "Sales in USD", "Market Share"),
    (14, 12200, 15),
    (20, 60000, 33),
    (18, 24400, 10),
    (22, 32000, 42)
]

for row in rows:
    sheet.append(row)


chart = BubbleChart()

# create data for plotting
xvalues = Reference(sheet, min_col=1,
                    min_row=2, max_row=5)

yvalues = Reference(sheet, min_col=2,
                    min_row=2, max_row=5)

size = Reference(sheet, min_col=3,
                 min_row=2, max_row=5)

# create a 1st series of data
series = Series(values=yvalues, xvalues=xvalues,
                zvalues=size, title="2013")

# add series data to the chart object
chart.series.append(series)

# set the title of the chart
chart.title = " BUBBLE-CHART "

# set the title of the x-axis
chart.x_axis.title = " X_AXIS "

# set the title of the y-axis
chart.y_axis.title = " Y_AXIS "

# add chart to the sheet
# the top-left corner of a chart
# is anchored to cell E2 .
sheet.add_chart(chart, "E2")

wb.save("BubbleChart.xlsx")
